import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def summarize_log_to_sheet(df, output_path="summary.xlsx", sheet_name="Summary"):
    """
    Create a summary sheet from the full log DataFrame using predefined fields.
    'Amount' will be filled from 'Grand Total' if missing.
    """
    # 🔧 Predefined fields
    summary_fields = [
        "Vehicle No", "Chassis No", "Transaction Date",
        "Amount", "Bank Ref No", "Vehicle Class", "NP Auth No", "Receipt No"
    ]
    
    # 🧪 Fill missing 'Amount' from 'Grand Total'
    df["Amount"] = df["Amount"].fillna(df["Grand Total"])
    # Fill NaN or 'NOT FOUND' in 'Amount' using 'Grand Total'
    df["Amount"] = df["Amount"].where(~df["Amount"].isin(["NOT FOUND", None, ""]), df["Grand Total"])

    
    # ✅ Validate fields
    missing = [field for field in summary_fields if field not in df.columns]
    if missing:
        raise ValueError(f"Missing fields in DataFrame: {missing}")
    
    # ✂️ Create summary
    summary_df = df[summary_fields].copy()
    
    # 🔢 Coerce numeric fields
    summary_df["Amount"] = pd.to_numeric(summary_df["Amount"], errors="coerce")
    # summary_df["Bank Ref No"] = pd.to_numeric(summary_df["Bank Ref No"], errors="coerce")
    
    # 📤 Save to Excel
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        summary_df.to_excel(writer, index=False, sheet_name=sheet_name)
    
    print(f"✅ Summary saved to '{output_path}'")

    # Apply formatting
    wb = load_workbook(output_path)
    ws = wb.active

    # Set font size to 9 for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(size=9)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_length + 2, 12)

    # 💾 Save formatted workbook
    wb.save(output_path)


# Load the full log Excel file
full_log_df = pd.read_excel("rto_receipts_log.xlsx")

# Run the summary function
summarize_log_to_sheet(full_log_df)


