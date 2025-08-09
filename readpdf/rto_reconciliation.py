# rto_reconciliation.py

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill

def apply_missing_highlight(ws):
    # Find the column index for "Receipt from RTO Portal"
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == "Receipt from RTO Portal":
            rto_col_letter = col[0].column_letter
            break
    else:
        return  # Column not found

    # Define red fill and font
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")

    # Apply conditional formatting rule
    formula = f'ISNUMBER(SEARCH("missing",{rto_col_letter}2))'
    ws.conditional_formatting.add(f"{rto_col_letter}2:{rto_col_letter}{ws.max_row}", FormulaRule(formula=[formula], fill=red_fill, font=red_font))

def normalize_date(date_str):
    # Try multiple formats, return only date part
    for fmt in ("%d-%m-%Y", "%d-%b-%Y", "%d-%m-%Y %H:%M:%S", "%d-%b-%Y %I:%M %p"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except:
            continue
    return None

def load_data(otp_path, summary_path):
    otp_df = pd.read_excel(otp_path)
    summary_df = pd.read_excel(summary_path)
    # Normalize dates
    
    otp_df['Norm Date'] = otp_df['Transaction Date'].apply(normalize_date)
    summary_df['Norm Date'] = summary_df['Transaction Date'].apply(normalize_date)
    return otp_df, summary_df

def match_transactions(otp_df, summary_df):
    results = []
    for _, row in otp_df.iterrows():
        matched_row = {}  # ✅ Always define this first

        match1 = summary_df[
            (summary_df['Vehicle No'] == row['Vehicle Reg. Number']) &
            (summary_df['Amount'] == row['RTO Amount']) &
            (summary_df['Norm Date'] == row['Norm Date'])
        ]
        match2 = summary_df[
            (summary_df['Chassis No'] == row['Chassis Number']) &
            (summary_df['Amount'] == row['RTO Amount']) &
            (summary_df['Norm Date'] == row['Norm Date'])
        ]

        if not match1.empty:
            status, scenario = "Matched", "Matched based on Vehicle No"
            matched_row = match1.iloc[0].to_dict()
        elif not match2.empty:
            status, scenario = "Matched", "Matched based on Chassis No"
            matched_row = match2.iloc[0].to_dict()
        else:
            status, scenario = "Missing", "None"

        results.append({
            **row.to_dict(), # original OTP row
            "Receipt from RTO Portal": "Available" if status == "Matched" else "Missing",
            "Match Scenario": scenario,
            **matched_row           # adds summary columns if matched
        })
    return pd.DataFrame(results)

def save_results(df, output_path):
    # Drop 'Norm Date' if present
    if 'Norm Date' in df.columns:
        df = df.drop(columns=['Norm Date'])

    # Reorder columns
    priority = ['Receipt from RTO Portal', 'Match Scenario']
    cols = df.columns.tolist()
    reordered = priority + [col for col in cols if col not in priority]
    df = df[reordered]

    # Save to Excel
    df.to_excel(output_path, index=False)

    # Apply formatting
    wb = load_workbook(output_path)
    ws = wb.active

    # Set font size to 9 for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(size=9)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_length + 2, 12)

    apply_missing_highlight(ws)    
    
    wb.save(output_path)

# Main execution
otp_df, summary_df = load_data("OTP_transaction_list.xlsx", "summary.xlsx")
result_df = match_transactions(otp_df, summary_df)
save_results(result_df, "reconciliation_result.xlsx")
