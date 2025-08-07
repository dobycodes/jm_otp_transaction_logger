from openpyxl import load_workbook
from datetime import datetime, timedelta

def is_recent_duplicate_transaction(
    excel_path,
    vehicle_number,
    chassis_number,
    payment_type,
    rto_amount,
    bank_amount
):
    import logging
    from openpyxl import load_workbook
    from datetime import datetime, timedelta

    # Normalize inputs
    vehicle_number = vehicle_number.strip() if vehicle_number else ""
    chassis_number = chassis_number.strip() if chassis_number else ""
    
    payment_type = str(payment_type).strip().lower() if payment_type else ""

    try:
        rto_amount = float(rto_amount)
        bank_amount = float(bank_amount)
    except ValueError:
        logging.warning("Amount conversion failed; skipping duplicate check.")
        return False

    if not vehicle_number and not chassis_number:
        logging.info("Skipping duplicate check: no identifiers provided.")
        return False

    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        threshold_date = datetime.today() - timedelta(days=4)

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                logged_time = datetime.strptime(str(row[0]), "%Y-%m-%d %H:%M:%S")
            except (ValueError, TypeError):
                continue

            if logged_time < threshold_date:
                continue

            logged_vehicle = str(row[1]).strip() if row[1] else ""
            logged_chassis = str(row[2]).strip() if row[2] else ""
            logged_payment = str(row[4]).strip().lower() if row[4] else ""
            try:
                logged_rto = float(row[5])
                logged_bank = float(row[6])
            except (ValueError, TypeError):
                continue

            identifiers_match = (
                (vehicle_number and logged_vehicle == vehicle_number) or
                (chassis_number and logged_chassis == chassis_number)
            )

            amounts_match = (
                logged_payment == payment_type and
                abs(logged_rto - rto_amount) < 0.01 and
                abs(logged_bank - bank_amount) < 0.01
            )

            if identifiers_match and amounts_match:
                logging.info(f"Duplicate found for {vehicle_number or chassis_number}")
                return True

        logging.info("No recent duplicate found.")
        return False

    except Exception as e:
        logging.error(f"Error checking for duplicates: {e}")
        return False
