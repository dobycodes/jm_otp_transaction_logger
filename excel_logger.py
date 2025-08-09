import os
from openpyxl import Workbook, load_workbook

def log_otp_to_excel(data, file_path="OTP_transaction_list.xlsx"):
    headers = [
        "Transaction Date", "Vehicle Reg. Number","Chassis Number", "Owner Name", "Payment Type",
        "RTO Amount", "Bank Amount", "OTP", "Employee Name",
        "Gmail Message ID", "Raw Email Body"
    ]

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "OTP Logs"
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    # Remove empty first row if present
    if ws.max_row >= 1 and all(cell.value is None for cell in ws[1]):
        ws.delete_rows(1)

    # Write headers to row 1 if missing
    if ws.max_row == 0 or [cell.value for cell in ws[1]] != headers:
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

    # Append data to next available row
    next_row = ws.max_row + 1
    ws.append([
        data["timestamp"].strftime("%Y-%m-%d %H:%M:%S") if data["timestamp"] else "",
        data.get("vehicle_reg",""),
        data.get("chassis_number", ""),
        data["owner_name"],
        data["payment_type"],
        data["rto_amount"],
        data["bank_amount"],
        data["otp"],
        data["employee_name"],
        data.get("gmail_id", ""),
        data.get("raw", "")
    ])

    wb.save(file_path)

