import gspread
import json
from openpyxl import load_workbook
from pathlib import Path
from oauth2client.service_account import ServiceAccountCredentials
from logger import setup_logger

# Initialize logger
logger = setup_logger(name="sheets_sync")

# 🔧 Load configuration
CONFIG_PATH = Path("config.json")
if not CONFIG_PATH.exists():
    logger.error(f"❌ config.json not found at: {CONFIG_PATH.resolve()}")
    raise FileNotFoundError("Missing config.json")

with open(CONFIG_PATH, "r") as f:
    config = json.load(f)

# Configurable paths
CREDENTIALS_PATH = Path(config["sheets_credentials_path"])
MASTER_SHEET_PATH = Path(config["transaction_log_excel_path"])
SHEET_ID = config.get("spreadsheet_id")  # ← google sheet ID

# Tab mappings: Excel → Google Sheets
TAB_MAPPING = config.get("tab_mapping", {})

def normalize(text):
    return str(text).strip().replace("\n", " ").replace("\r", "")

def push_tab(excel_tab, sheet_tab):
    try:
        # Authenticate
        creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_PATH, [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ])
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(SHEET_ID)

        # Load Excel workbook
        wb = load_workbook(MASTER_SHEET_PATH, data_only=True)
        if excel_tab not in wb.sheetnames:
            logger.warning(f"⚠️ Excel tab '{excel_tab}' not found in workbook. Skipping sync.")
            return

        ws = wb[excel_tab]
        excel_headers = [normalize(cell.value) for cell in ws[1]]
        logger.debug(f"🔍 Headers from '{excel_tab}': {excel_headers}")

        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cleaned_row = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(cleaned_row):
                rows.append(cleaned_row)
            else:
                logger.debug(f"⏭️ Skipped empty row {i} in '{excel_tab}'")

        if config.get("dry_run"):
            logger.info(f"🧪 Dry run: would sync {len(rows)} rows to '{sheet_tab}' from '{excel_tab}'")
            return

        # Ensure tab exists
        try:
            sheet = spreadsheet.worksheet(sheet_tab)
        except gspread.exceptions.WorksheetNotFound:
            logger.warning(f"⚠️ Tab '{sheet_tab}' not found. Creating new worksheet...")
            sheet = spreadsheet.add_worksheet(title=sheet_tab, rows="1000", cols="50")
            logger.info(f"✅ Created new tab: '{sheet_tab}'")

        sheet.clear()
        sheet.append_row(excel_headers)
        if rows:
            sheet.append_rows(rows, value_input_option="USER_ENTERED")
            logger.info(f"✅ Synced {len(rows)} rows to '{sheet_tab}' from '{excel_tab}'")
        else:
            logger.info(f"ℹ️ No non-empty rows to sync for '{excel_tab}'")

    except Exception as e:
        logger.error(f"❌ Sync failed for '{excel_tab}': {e}")

def push_to_google_sheet():
    for excel_tab, sheet_tab in TAB_MAPPING.items():
        push_tab(excel_tab, sheet_tab)

if __name__ == "__main__":
    push_to_google_sheet()
