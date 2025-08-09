import json
from pathlib import Path
from openpyxl import load_workbook
from google.oauth2 import service_account
from googleapiclient.discovery import build
from logger import setup_logger

logger = setup_logger(name="sheets_downsync")

# 🔧 Load configuration
CONFIG_PATH = Path("config.json")
if not CONFIG_PATH.exists():
    logger.error(f"❌ config.json not found at: {CONFIG_PATH.resolve()}")
    raise FileNotFoundError("Missing config.json")

with open(CONFIG_PATH, "r") as f:
    config = json.load(f)

CREDENTIALS_PATH = Path(config["sheets_credentials_path"])
MASTER_SHEET_PATH = Path(config["transaction_log_excel_path"])
SHEET_ID = config.get("spreadsheet_id")  # ← google sheet ID

TAB_MAPPING = config.get("tab_mapping", {})

# Function to refresh transaction types from an Excel tab and write to JSON
def refresh_transaction_types(excel_path: Path, tab_name: str, json_path: Path):
    if not excel_path.exists():
        logger.error(f"❌ Excel file not found: {excel_path.resolve()}")
        return

    try:
        wb = load_workbook(excel_path)
        if tab_name not in wb.sheetnames:
            logger.error(f"❌ Tab '{tab_name}' not found in workbook.")
            return
        ws = wb[tab_name]
    except Exception as e:
        logger.error(f"❌ Failed to load workbook/tab: {e}")
        return

    types = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        if row and row[0]:  # Assuming first column is 'Type'
            types.append(str(row[0]).strip())

    if not types:
        logger.warning(f"⚠️ No transaction types found in tab '{tab_name}'")
        return

    try:
        with open(json_path, "w") as f:
            json.dump({"payment_types": types}, f, indent=2)
        logger.info(f"✅ Refreshed '{json_path.name}' with {len(types)} types.")
    except Exception as e:
        logger.error(f"❌ Failed to write JSON: {e}")

# Function to ensure the Excel file exists and has the required tabs
def ensure_excel_exists(path: Path, tab_names: list[str]):
    from openpyxl import Workbook

    if not path.exists():
        logger.warning(f"⚠️ Excel file '{path.name}' not found. Creating new workbook...")
        wb = Workbook()
        wb.remove(wb.active)

        for tab_name in tab_names:
            wb.create_sheet(title=tab_name)
            logger.info(f"🆕 Created tab '{tab_name}' in new workbook.")

        wb.save(path)
        logger.info(f"✅ Created new Excel file with tabs: {', '.join(tab_names)}")

# Function to pull data from a Google Sheets tab and write it to an Excel tab
def pull_tab(sheet_tab, local_tab):
    
    if not CREDENTIALS_PATH.exists():
        logger.error(f"❌ Credential file not found: {CREDENTIALS_PATH.resolve()}")
        return

    ensure_excel_exists(MASTER_SHEET_PATH, list(TAB_MAPPING.values()))

    try:
        creds = service_account.Credentials.from_service_account_file(
            str(CREDENTIALS_PATH),
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        service = build("sheets", "v4", credentials=creds)
    except Exception as e:
        logger.error(f"❌ Failed to initialize Google Sheets service: {e}")
        return

    try:
        result = service.spreadsheets().values().get(
            spreadsheetId=SHEET_ID,
            range=sheet_tab
        ).execute()
        rows = result.get("values", [])
    except Exception as e:
        logger.error(f"❌ Failed to fetch data from tab '{sheet_tab}': {e}")
        return

    if not rows:
        logger.warning(f"⚠️ No data found in Sheets tab '{sheet_tab}'")
        return

    try:
        wb = load_workbook(MASTER_SHEET_PATH)
        if local_tab not in wb.sheetnames:
            logger.warning(f"⚠️ Excel tab '{local_tab}' not found in workbook. Creating it...")
            wb.create_sheet(local_tab)
        ws = wb[local_tab]
    except Exception as e:
        logger.error(f"❌ Failed to access workbook or tab '{local_tab}': {e}")
        return

    if ws.max_row > 1:
        ws.delete_rows(1, ws.max_row)


    for i, row in enumerate(rows, start=1):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j).value = value

    wb.save(MASTER_SHEET_PATH)
    logger.info(f"✅ Reverse sync completed for '{local_tab}'. {len(rows)} rows copied.")

# Function to pull all tabs from Google Sheets to local Excel
def pull_from_google_sheet():
    logger.info("🔄 Starting full reverse sync from Google Sheets to local Excel...")
    for sheet_tab, local_tab in TAB_MAPPING.items():
        pull_tab(sheet_tab, local_tab)
